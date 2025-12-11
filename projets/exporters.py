# exporters.py
import pandas as pd
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExcelExporter:
    def __init__(self, projet, lots):
        self.projet = projet
        self.lots = lots
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Supprimer la feuille par défaut
        
    def apply_styles(self, ws):
        """Applique des styles professionnels à une feuille"""
        # Styles pour les en-têtes
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Styles pour les données
        data_font = Font(size=10)
        currency_format = '#,##0.00'
        
        # Appliquer aux en-têtes
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        # Ajuster les largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    def create_summary_sheet(self):
        """Crée la feuille de résumé"""
        ws = self.wb.create_sheet(title="Résumé")
        
        # Titre
        ws['A1'] = f"BORDEREAU DE PRIX - {self.projet.nom}"
        ws['A1'].font = Font(size=14, bold=True, color="4ECDC4")
        
        # Informations du projet
        ws['A3'] = "Informations du projet"
        ws['A3'].font = Font(bold=True)
        
        data = [
            ["Numéro", self.projet.numero],
            ["Nom", self.projet.nom],
            ["Date de génération", datetime.now().strftime("%d/%m/%Y %H:%M")],
            ["", ""],
            ["Résumé financier", ""],
            ["Total lots", len(self.lots)],
            ["Montant total HT", float(self.projet.montant_total())],
            ["Montant total TTC", float(self.projet.montant_total()) * 1.2],  # Exemple TVA 20%
        ]
        
        for i, row in enumerate(data, start=4):
            ws[f'A{i}'] = row[0]
            ws[f'B{i}'] = row[1]
            if isinstance(row[1], (int, float)):
                ws[f'B{i}'].number_format = '#,##0.00'
        
        return ws
    
    def create_lot_sheet(self, lot):
        """Crée une feuille pour un lot"""
        ws = self.wb.create_sheet(title=lot.nom[:31])  # Excel limite à 31 caractères
        
        # En-tête du lot
        ws['A1'] = f"LOT: {lot.nom}"
        ws['A1'].font = Font(size=12, bold=True, color="2C3E50")
        
        if lot.description:
            ws['A2'] = lot.description
        
        # En-têtes du tableau
        headers = ["N°", "Désignation", "Unité", "Quantité", "PU (MAD)", "Montant (MAD)"]
        ws.append(headers)
        
        # Données
        for ligne in lot.lignes.all():
            indentation = "    " * ligne.level
            row = [
                ligne.numero or "",
                indentation + ligne.designation,
                ligne.unite or "",
                ligne.quantite or 0,
                ligne.prix_unitaire or 0,
                ligne.montant or 0
            ]
            ws.append(row)
        
        # Total
        ws.append(["", "", "", "", "Total HT:", lot.montant_total_ht])
        ws.append(["", "", "", "", "Total TTC:", lot.montant_total_ttc])  # Exemple TVA
        
        # Appliquer les styles
        self.apply_styles(ws)
        
        # Format monétaire
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                if cell.column in [5, 6]:  # Colonnes PU et Montant
                    cell.number_format = '#,##0.00'
        
        return ws
    
    def export(self):
        """Génère l'export Excel"""
        # Créer la feuille de résumé
        self.create_summary_sheet()
        
        # Créer une feuille par lot
        for lot in self.lots:
            self.create_lot_sheet(lot)
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        # Retourner la réponse HTTP
        filename = f"bordereau_{self.projet.numero}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response